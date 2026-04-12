import flet as ft
import os
import sys
import threading
import time
import re
import subprocess
import json
from datetime import datetime
import uiautomator2 as u2
from nuitka_compat import ensure_nuitka_compatibility
from adb_utils import ADBUtils as adb_utils
from config_manager import config_manager
from logger_manager import get_logger

# 确保 Nuitka 兼容性
ensure_nuitka_compatibility()

logger = get_logger("ADBTools.Flet")

# --- Helper Classes ---

class FletLogger:
    def __init__(self, log_view):
        self.log_view = log_view
        self._buffer = []

    def write(self, message):
        if not message or not message.strip():
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_msg = f"[{timestamp}] {message.strip()}\n"
        self._buffer.append(formatted_msg)
        if len(self._buffer) > 500:
            self._buffer.pop(0)
        
        self.log_view.value = "".join(self._buffer)
        try:
            self.log_view.update()
        except:
            pass

    def flush(self):
        pass

class FileBrowser(ft.Container):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.current_path = "/sdcard"
        self.files = []
        self.loading = False
        self.expand = True
        self.content = self._build_ui()

    def _build_ui(self):
        self.file_list = ft.ListView(expand=True, spacing=5)
        self.path_text = ft.Text(self.current_path, weight="bold", size=16)
        
        return ft.Column([
            ft.Row([
                ft.IconButton(ft.Icons.ARROW_BACK, on_click=self.go_back, tooltip="返回上一级"),
                self.path_text,
                ft.IconButton(ft.Icons.REFRESH, on_click=self.refresh, tooltip="刷新目录"),
                ft.IconButton(ft.Icons.UPLOAD_FILE, on_click=self.pick_upload_file, tooltip="上传文件到此目录"),
            ]),
            ft.Divider(),
            ft.Container(content=self.file_list, expand=True, height=400),
        ], expand=True)

    def pick_upload_file(self, _):
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            paths = filedialog.askopenfilenames(title="选择要上传的文件")
            root.destroy()
            if paths:
                self.upload_files(list(paths))
        except Exception as e:
            self.app.log(f"选择文件失败: {e}")

    def upload_files(self, local_paths):
        if not self.app.selected_device: return
        for local_path in local_paths:
            filename = os.path.basename(local_path)
            remote_path = os.path.join(self.current_path, filename).replace("\\", "/")
            self.app.log(f"正在上传: {filename} -> {remote_path}")
            def task(lp=local_path, rp=remote_path, fn=filename):
                res = adb_utils.run_adb_command(f"push \"{lp}\" \"{rp}\"", self.app.selected_device)
                if res.returncode == 0:
                    self.app.log(f"上传成功: {fn}")
                    self.refresh()
                else:
                    self.app.log(f"上传失败 ({fn}): {res.stderr}")
            threading.Thread(target=task).start()

    def go_back(self, _):
        if self.current_path == "/": return
        self.current_path = os.path.dirname(self.current_path.rstrip("/"))
        if not self.current_path: self.current_path = "/"
        self.refresh()

    def refresh(self, _=None):
        if not self.app.selected_device: return
        self.app.log(f"正在读取目录: {self.current_path}")
        def task():
            res = adb_utils.run_adb_command(f"shell ls -F {self.current_path}", self.app.selected_device)
            if res.returncode == 0:
                lines = res.stdout.strip().split("\n")
                self.files = []
                for line in lines:
                    if not line.strip(): continue
                    is_dir = line.endswith("/")
                    name = line.rstrip("/")
                    self.files.append({"name": name, "is_dir": is_dir})
                
                self.update_ui()
            else:
                self.app.log(f"读取目录失败: {res.stderr}")
        threading.Thread(target=task).start()

    def update_ui(self):
        self.file_list.controls.clear()
        self.path_text.value = self.current_path
        for f in self.files:
            icon = ft.Icons.FOLDER if f["is_dir"] else ft.Icons.INSERT_DRIVE_FILE
            
            trailing = None
            if not f["is_dir"]:
                trailing = ft.Row([
                    ft.IconButton(ft.Icons.DOWNLOAD, tooltip="拉取到本地", on_click=lambda _, name=f["name"]: self.download_file(name)),
                    ft.IconButton(ft.Icons.DELETE, tooltip="从设备删除", icon_color="red400", on_click=lambda _, name=f["name"]: self.delete_file(name)),
                ], tight=True)
            else:
                trailing = ft.IconButton(ft.Icons.DELETE, tooltip="删除目录", icon_color="red400", on_click=lambda _, name=f["name"]: self.delete_file(name))

            self.file_list.controls.append(
                ft.ListTile(
                    leading=ft.Icon(icon),
                    title=ft.Text(f["name"]),
                    trailing=trailing,
                    on_click=lambda _, name=f["name"], is_dir=f["is_dir"]: self.on_file_click(name, is_dir)
                )
            )
        self.update()

    def download_file(self, name):
        remote_path = os.path.join(self.current_path, name).replace("\\", "/")
        self.app.log(f"正在拉取文件: {remote_path}")
        def task():
            res = adb_utils.run_adb_command(f"pull \"{remote_path}\" .", self.app.selected_device)
            if res.returncode == 0:
                self.app.log(f"成功拉取到当前目录: {name}")
            else:
                self.app.log(f"拉取失败: {res.stderr}")
        threading.Thread(target=task).start()

    def delete_file(self, name):
        remote_path = os.path.join(self.current_path, name).replace("\\", "/")
        self.app.log(f"正在删除: {remote_path}")
        def task():
            res = adb_utils.run_adb_command(f"shell rm -rf \"{remote_path}\"", self.app.selected_device)
            if res.returncode == 0:
                self.app.log(f"已从设备删除: {name}")
                self.refresh()
            else:
                self.app.log(f"删除失败: {res.stderr}")
        threading.Thread(target=task).start()

    def on_file_click(self, name, is_dir):
        if is_dir:
            self.current_path = os.path.join(self.current_path, name).replace("\\", "/")
            self.refresh()
        else:
            self.app.log(f"选中文件: {name} (此处可增加下载/删除逻辑)")

# --- Main App ---

class ADBToolsApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = f"ADBTools Flet - v{config_manager.get_version()}"
        self.page.theme_mode = ft.ThemeMode.DARK
        self.page.window.width = 1200
        self.page.window.height = 900
        
        # 状态变量
        self.selected_device = None
        self.devices = []
        self.d = None # uiautomator2 instance
        self.u2_connecting = False
        self.connection_mode = "u2" # Default to u2 as in PyQt
        self.pkg_input = ft.Ref[ft.TextField]()
        self.text_input = ft.Ref[ft.TextField]()
        self.filter_input_ref = ft.Ref[ft.TextField]()
        self.excel_path = None
        self.file_picker_mode = "install" # "install" or "verify"
        self.operation_history = []
        self.all_packages = [] # Store for filtering
        self.pull_log_process = None # Current logcat process
        self.pull_log_active = False # Flag for logcat loop
        
        # 特殊包配置 (从 ADBBatchInstallThread 移植)
        self.special_packages_config = config_manager.get("batch_install.special_packages", {
            "com.saicmotor.voiceservice": {
                "delete_before_push": False,
                "description": "voiceservice包，只push不删除"
            },
            "com.saicmotor.adapterservice": {
                "delete_before_push": True,
                "description": "adapterservice包，先删除再push"
            }
        })
        
        # UI 组件
        self.log_view = ft.TextField(
            multiline=True,
            read_only=True,
            expand=True,
            text_size=12,
            text_style=ft.TextStyle(font_family="Consolas"),
            bgcolor="black",
            color="green300",
            border_color="grey800",
        )
        self.flet_logger = FletLogger(self.log_view)
        
        self.device_dropdown = ft.Dropdown(
            label="选择设备",
            on_select=self.on_device_selected,
            expand=True,
            hint_text="请先连接并刷新设备"
        )
        
        self.main_container = ft.Container(expand=True, padding=15)
        self.sidebar = None
        
        self.setup_ui()
        self.refresh_devices()
        
        # 启动时静默检查更新
        threading.Thread(target=self.check_software_update, args=(None, True), daemon=True).start()

    def setup_ui(self):
        # 头部
        header = ft.Container(
            content=ft.Row([
                ft.Row([
                    ft.Icon(ft.Icons.TERMINAL, color="blue400", size=30),
                    ft.Text("ADBTools", size=24, weight="bold"),
                ]),
                ft.Row([
                    ft.Checkbox(label="U2", value=True, on_change=self.on_mode_switch_changed),
                    ft.Container(self.device_dropdown, width=300),
                    ft.Button("刷新设备", icon=ft.Icons.REFRESH, on_click=lambda _: self.refresh_devices()),
                    ft.Button("断开设备", icon=ft.Icons.LINK_OFF, on_click=self.disconnect_device, color="red400"),
                ], spacing=10),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            padding=ft.Padding.symmetric(horizontal=20, vertical=10),
            bgcolor=ft.Colors.SURFACE_CONTAINER,
            border_radius=10,
        )

        # 侧边栏
        self.sidebar = ft.NavigationRail(
            selected_index=0,
            label_type=ft.NavigationRailLabelType.ALL,
            min_width=100,
            min_extended_width=160,
            destinations=[
                ft.NavigationRailDestination(icon=ft.Icons.HOME_ROUNDED, label="ADB工具"),
                ft.NavigationRailDestination(icon=ft.Icons.APPS, label="大通项目"),
                ft.NavigationRailDestination(icon=ft.Icons.CAR_REPAIR, label="CR项目"),
                ft.NavigationRailDestination(icon=ft.Icons.HISTORY, label="Pull Log"),
                ft.NavigationRailDestination(icon=ft.Icons.LANGUAGE, label="网联版项目"),
                ft.NavigationRailDestination(icon=ft.Icons.VOICE_CHAT, label="语音相关"),
                ft.NavigationRailDestination(icon=ft.Icons.CHECK_CIRCLE, label="集成版本检查"),
                ft.NavigationRailDestination(icon=ft.Icons.SETTINGS, label="设置"),
            ],
            on_change=self.on_nav_change,
        )

        # 主布局
        self.page.add(
            ft.Column([
                header,
                ft.Row([
                    self.sidebar,
                    ft.VerticalDivider(width=1),
                    ft.Column([
                        self.main_container,
                        ft.Divider(height=1, color="grey800"),
                        ft.Row([
                            ft.Text(" 控制台输出", size=12, weight="bold", color="grey500"),
                            ft.Row([
                                ft.Button("清空", icon=ft.Icons.DELETE_SWEEP, on_click=lambda _: self.clear_flet_logs()),
                                ft.Button("重新初始化U2", icon=ft.Icons.REFRESH, on_click=self.reinit_u2, color="orange400"),
                            ]),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Container(
                            content=self.log_view,
                            height=200,
                            padding=5,
                            bgcolor="black",
                        ),
                        ft.Text("By wk WX:2315313745", size=10, color="grey500"),
                    ], expand=True)
                ], expand=True)
            ], expand=True, spacing=0)
        )
        
        self.show_home()

    def on_mode_switch_changed(self, e):
        self.connection_mode = "u2" if e.control.value else "adb"
        self.log(f"切换连接模式为: {self.connection_mode.upper()}")
        if self.connection_mode == "u2":
            self.connect_u2()

    def connect_u2(self):
        if not self.selected_device: return
        if self.u2_connecting: return
        self.log(f"正在尝试 U2 连接: {self.selected_device}")
        def task():
            self.u2_connecting = True
            try:
                # 尝试连接
                self.d = u2.connect(self.selected_device)
                # 简单测试连接是否有效
                info = self.d.info
                self.log(f"U2 连接成功! 设备型号: {info.get('productName', '未知')}")
            except Exception as e:
                self.log(f"U2 连接失败: {e}")
                self.d = None
            finally:
                self.u2_connecting = False
                self.page.update()
        threading.Thread(target=task).start()

    def on_nav_change(self, e):
        idx = e.control.selected_index
        if idx == 0: self.show_home()
        elif idx == 1: self.show_datong()
        elif idx == 2: self.show_cr_project()
        elif idx == 3: self.show_pull_log()
        elif idx == 4: self.show_internet_project()
        elif idx == 5: self.show_voice_related()
        elif idx == 6: self.show_version_check()
        elif idx == 7: self.show_settings()
        self.page.update()

    def clear_flet_logs(self):
        self.flet_logger._buffer = []
        self.log_view.value = ""
        self.page.update()

    def log(self, message):
        self.flet_logger.write(message)
        if "执行:" in message or "正在" in message:
            self.operation_history.append(message)
            if len(self.operation_history) > 50:
                self.operation_history.pop(0)

    def on_device_selected(self, e):
        selected = getattr(e, "data", None)
        if not selected and hasattr(e, "control"):
            selected = getattr(e.control, "value", None)
        self.selected_device = selected
        self.log(f"当前选中设备: {self.selected_device}")
        if self.connection_mode == "u2":
            self.connect_u2()

    def disconnect_device(self, _):
        if not self.selected_device: return
        self.log(f"正在断开设备: {self.selected_device}")
        def task():
            adb_utils.run_adb_command(f"disconnect {self.selected_device}")
            self.refresh_devices()
        threading.Thread(target=task).start()

    def refresh_devices(self):
        self.log("正在刷新 ADB 设备...")
        def task():
            ids = adb_utils.get_device_list()
            self.devices = ids
            self.device_dropdown.options = [ft.dropdown.Option(d) for d in ids]
            if ids:
                if not self.selected_device or self.selected_device not in ids:
                    self.selected_device = ids[0]
                    self.device_dropdown.value = self.selected_device
                self.log(f"成功刷新，在线设备: {len(ids)}")
            else:
                self.selected_device = None
                self.device_dropdown.value = None
                self.log("未发现在线设备")
            try:
                self.page.update()
            except:
                pass
        threading.Thread(target=task).start()

    def create_action_button(self, text, icon, on_click, color="blue400"):
        if isinstance(icon, str):
            icon = getattr(ft.Icons, icon.strip().upper(), ft.Icons.HELP_OUTLINE)
        return ft.Button(
            content=ft.Row([ft.Icon(icon, size=20, color=color), ft.Text(text)], tight=True),
            on_click=on_click,
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
        )

    # --- 页面渲染 ---

    def show_home(self):
        self.main_container.content = ft.Column([
            ft.Row([
                self.create_action_button("获取Root权限", ft.Icons.SECURITY, self.adb_root),
                self.create_action_button("重启设备", ft.Icons.RESTART_ALT, self.reboot_device),
            ], spacing=10),
            ft.TextField(
                label="请输入要过滤的包名", 
                hint_text="请输入要过滤的包名", 
                on_change=self.filter_packages,
                prefix_icon=ft.Icons.SEARCH,
                ref=self.filter_input_ref
            ),
            self.create_action_button("列出所有包名 findstr ↑", ft.Icons.LIST, lambda _: self.list_packages(self.filter_input_ref.current.value)),
            ft.Row([
                self.create_action_button("安装应用", ft.Icons.FILE_UPLOAD, self.install_apk_dialog),
                self.create_action_button("文件管理", ft.Icons.FOLDER_SHARED, self.show_file_browser),
            ], spacing=10),
            ft.Row([
                self.create_action_button("卸载应用", ft.Icons.DELETE_OUTLINE, self.app_uninstall, color="red400"),
                self.create_action_button("启动应用", ft.Icons.PLAY_ARROW, self.app_start),
            ], spacing=10),
            self.create_action_button("重启ADB服务", ft.Icons.REFRESH, self.restart_adb_service),
            ft.Divider(),
            ft.Row([
                self.create_action_button("获取应用版本号", ft.Icons.INFO, self.get_running_app_info),
                self.create_action_button("获取apk文件的包名", ft.Icons.SEARCH, self.aapt_getpackage_name_dilog),
            ], spacing=10),
            self.create_action_button("adb shell pm path", ft.Icons.FOLDER, self.view_app_path),
            ft.Row([
                self.create_action_button("Input text", ft.Icons.TEXT_FIELDS, self.show_input_dialog),
                self.create_action_button("Screen", ft.Icons.SCREENSHOT, self.take_screenshot),
            ], spacing=10),
            self.create_action_button("获取包名和活动页名", ft.Icons.GPS_FIXED, self.get_current_app),
            ft.Row([
                self.create_action_button("Clear App 缓存", ft.Icons.CLEANING_SERVICES, self.app_clear_data),
                self.create_action_button("Close App", ft.Icons.STOP, self.app_force_stop, color="red400"),
            ], spacing=10),
            ft.TextField(label="包名 (Package Name)", hint_text="输入包名进行快捷操作", ref=self.pkg_input),
        ], scroll=ft.ScrollMode.AUTO)

    def restart_adb_service(self, _):
        self.log("正在重启 ADB 服务...")
        def task():
            subprocess.run(["adb", "kill-server"])
            subprocess.run(["adb", "start-server"])
            self.log("ADB 服务已重启")
            self.refresh_devices()
        threading.Thread(target=task).start()

    def list_packages(self, filter_text):
        if not self.check_device(): return
        self.log(f"正在获取包名列表 (过滤: {filter_text})...")
        def task():
            # 在 Android shell 中，grep 几乎总是可用的
            # 报错可能是因为整个命令字符串在 Windows cmd/powershell 中被误解析
            # 确保 | grep 部分是作为 adb shell 的一部分传递的
            cmd = f"shell \"pm list packages | grep '{filter_text}'\"" if filter_text else "shell pm list packages"
            res = adb_utils.run_adb_command(cmd, self.selected_device)
            
            if res.returncode == 0:
                pkgs = res.stdout.strip().split("\n")
                self.all_packages = [p.replace("package:", "").strip() for p in pkgs if p.strip()]
                self.log(f"找到 {len(self.all_packages)} 个包:")
                for p in self.all_packages:
                    self.log(p)
            else:
                self.log(f"获取包名失败: {res.stderr}")
        threading.Thread(target=task).start()

    def filter_packages(self, e):
        # 此处可以实现实时过滤逻辑，目前仅作为占位符或简单日志
        pass

    def show_file_browser(self, _):
        self.main_container.content = ft.Column([
            ft.Text("设备文件浏览器", size=18, weight="bold"),
            FileBrowser(self),
        ], scroll=ft.ScrollMode.AUTO)
        self.page.update()

    def _tk_pick_file(self, title, filetypes):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        root.destroy()
        return path

    def _tk_pick_files(self, title, filetypes):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        paths = filedialog.askopenfilenames(title=title, filetypes=filetypes)
        root.destroy()
        return list(paths)

    def _tk_pick_dir(self, title):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askdirectory(title=title)
        root.destroy()
        return path

    def install_apk_dialog(self, _):
        path = self._tk_pick_file("选择应用安装包", [("APK Files", "*.apk"), ("All Files", "*.*")])
        if not path:
            return
        self.log(f"准备安装 APK: {path}")
        self.run_generic_cmd(f"install -r \"{path}\"")

    def aapt_getpackage_name_dilog(self, _):
        path = self._tk_pick_file("选择 APK 文件", [("APK Files", "*.apk"), ("All Files", "*.*")])
        if not path:
            return
        self.log(f"正在获取 APK 包名: {path}")
        def task():
            pkg = adb_utils.aapt_get_package_name(path)
            self.log(f"识别到包名: {pkg}")
            if "失败" not in pkg and self.pkg_input.current:
                self.pkg_input.current.value = pkg
                self.page.update()
        threading.Thread(target=task).start()

    def show_datong(self):
        self.main_container.content = ft.Column([
            ft.Row([
                self.create_action_button("拉起工程模式", ft.Icons.BUILD, lambda _: self.run_generic_cmd("shell am start -n com.zhonghuan.factory/.MainActivity")),
                self.create_action_button("批量安装APK文件", ft.Icons.LIBRARY_ADD, self.pick_install_dir),
            ], spacing=10),
            ft.Row([
                self.create_action_button("启用verity校验 (adb enable-verity)", ft.Icons.VERIFIED_USER, lambda _: self.run_generic_cmd("enable-verity")),
                self.create_action_button("验证批量推包版本号", ft.Icons.VERIFIED, self.pick_verify_dir),
            ], spacing=10),
            ft.Row([
                self.create_action_button("禁用verity校验 (adb disable-verity)", ft.Icons.NO_ENCRYPTION, lambda _: self.run_generic_cmd("disable-verity")),
                self.create_action_button("一键输入密码 (Kfs73p940a)", ft.Icons.PASSWORD, lambda _: self.run_generic_cmd("shell input text Kfs73p940a")),
            ], spacing=10),
            ft.Row([
                self.create_action_button("设置设备日期时间", ft.Icons.ACCESS_TIME, self.sync_device_time),
                self.create_action_button("打开泰维地图工程模式", ft.Icons.MAP, lambda _: self.run_generic_cmd("shell am start -n com.telenav.app.denali.pnd/com.telenav.app.android.pnd.view.debug.EngineeringModeActivity")),
            ], spacing=10),
        ], scroll=ft.ScrollMode.AUTO)

    def show_cr_project(self):
        self.main_container.content = ft.Column([
            self.create_action_button("AS33_CR进入工程模式", ft.Icons.SETTINGS_SUGGEST, lambda _: self.run_generic_cmd("shell am start -n com.saicmotor.diag/.MainActivity")),
            self.create_action_button("AS33R国项目打开工程模式", ft.Icons.SETTINGS_APPLICATIONS, lambda _: self.run_generic_cmd("shell am start -n com.saicmotor.diag/.ui.main.MainActivity")),
        ], scroll=ft.ScrollMode.AUTO)

    def show_pull_log(self):
        self.log_path_input = ft.TextField(
            label="日志保存路径",
            value=os.getcwd(),
            expand=True,
            suffix=ft.IconButton(ft.Icons.FOLDER_OPEN, on_click=lambda _: self.pick_log_save_path()),
        )
        self.pull_log_btn = self.create_action_button(
            "停止拉取日志" if self.pull_log_active else "开始持续拉取日志",
            ft.Icons.STOP if self.pull_log_active else ft.Icons.PLAY_ARROW,
            self.toggle_pull_logs,
            color="red400" if self.pull_log_active else "blue400"
        )
        self.main_container.content = ft.Column([
            ft.Text("日志路径：", size=16),
            ft.Row([
                self.log_path_input,
                self.create_action_button("浏览", ft.Icons.FOLDER_OPEN, lambda _: self.pick_log_save_path()),
                self.create_action_button("打开目录", ft.Icons.OPEN_IN_NEW, self.open_local_log_dir),
            ], spacing=10),
            self.pull_log_btn,
        ], scroll=ft.ScrollMode.AUTO)

    def toggle_pull_logs(self, _):
        if self.pull_log_active:
            self.stop_pull_logs()
        else:
            self.start_pull_logs()

    def start_pull_logs(self):
        if not self.check_device(): return
        save_dir = self.log_path_input.value if hasattr(self, 'log_path_input') else os.getcwd()
        filename = os.path.join(save_dir, f"logcat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        try:
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            # 使用 subprocess.Popen 启动持续进程
            cmd = ["adb", "-s", self.selected_device, "logcat", "-v", "time"]
            self.log(f"开始持续拉取日志到: {filename}")
            
            self.pull_log_active = True
            self.pull_log_file = open(filename, "w", encoding="utf-8")
            self.pull_log_process = subprocess.Popen(
                cmd, 
                stdout=self.pull_log_file, 
                stderr=subprocess.STDOUT,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            # 更新按钮状态
            self.pull_log_btn.text = "停止拉取日志"
            self.pull_log_btn.icon = ft.Icons.STOP
            self.pull_log_btn.color = "red400"
            self.page.update()
            
        except Exception as e:
            self.log(f"启动日志拉取失败: {e}")
            self.pull_log_active = False

    def stop_pull_logs(self):
        if self.pull_log_process:
            try:
                self.pull_log_process.terminate()
                self.pull_log_process.wait(timeout=2)
            except:
                self.pull_log_process.kill()
            
            self.pull_log_process = None
            
        if hasattr(self, 'pull_log_file') and self.pull_log_file:
            self.pull_log_file.close()
            self.pull_log_file = None
            
        self.pull_log_active = False
        self.log("已停止日志拉取")
        
        # 更新按钮状态
        if hasattr(self, 'pull_log_btn'):
            self.pull_log_btn.text = "开始持续拉取日志"
            self.pull_log_btn.icon = ft.Icons.PLAY_ARROW
            self.pull_log_btn.color = "blue400"
            self.page.update()

    def pick_log_save_path(self):
        path = self._tk_pick_dir("选择日志保存路径")
        if not path:
            return
        self.log(f"已选择日志保存路径: {path}")
        if hasattr(self, 'log_path_input'):
            self.log_path_input.value = path
            self.page.update()

    def show_internet_project(self):
        self.main_container.content = ft.Column([
            self.create_action_button("调起延峰升级页面", ft.Icons.SYSTEM_UPDATE_ALT, lambda _: self.run_generic_cmd("shell am start -n com.yfve.usbupdate/.MainActivity")),
            self.create_action_button("跳过升级时的电源模式与挡位限制", ft.Icons.BOLT, lambda _: self.run_generic_cmd("shell setprop vr.power.limit.skip 1")),
            self.create_action_button("调起资源升级页面", ft.Icons.CLOUD_DOWNLOAD, lambda _: self.run_generic_cmd("shell am start -n com.saicmotor.update/.view.MainActivity")),
            self.create_action_button("进入工程模式", ft.Icons.DEVELOPER_MODE, lambda _: self.run_generic_cmd("shell am start -n com.saicmotor.hmi.engmode/.MainActivity")),
        ], scroll=ft.ScrollMode.AUTO)

    def show_voice_related(self):
        self.voice_path_input = ft.TextField(
            label="车机端录音路径", 
            value="/vr/speech/assistant/files/tmp/audioDump",
            hint_text="输入车机端录音文件路径"
        )
        self.vr_key_combo = ft.Dropdown(
            value="287",
            options=[ft.dropdown.Option("287"), ft.dropdown.Option("231")],
            width=100
        )
        self.main_container.content = ft.Column([
            ft.Text("车机语音识别原始录音文件操作", size=18, weight="bold"),
            ft.Row([
                self.create_action_button("车机开始录音", ft.Icons.MIC, self.voice_start_record),
                self.create_action_button("停止录音", ft.Icons.MIC_OFF, self.voice_stop_record, color="red400"),
            ], spacing=10),
            self.voice_path_input,
            ft.Row([
                self.create_action_button("拉取录音文件", ft.Icons.FILE_DOWNLOAD, self.voice_pull_record),
                self.create_action_button("删除历史录音文件", ft.Icons.DELETE_FOREVER, self.voice_remove_record, color="red400"),
            ], spacing=10),
            ft.Divider(),
            ft.Text("VR测试", size=18, weight="bold"),
            ft.Row([
                self.create_action_button("切换VR环境", ft.Icons.SWAP_HORIZ, lambda _: self.run_generic_cmd("shell am broadcast -a com.saicmotor.vr.SWITCH_ENV")),
                self.create_action_button("VR网络环境检查", ft.Icons.NETWORK_CHECK, lambda _: self.run_generic_cmd("shell ping -c 4 8.8.8.8")),
            ], spacing=10),
            ft.Row([
                self.create_action_button("调起VR", ft.Icons.VOICE_OVER_OFF, lambda _: self.run_generic_cmd(f"shell input keyevent {self.vr_key_combo.value}")),
                self.vr_key_combo,
            ], spacing=10),
            self.create_action_button("设置语音超时时间为10000", ft.Icons.TIMER, self.show_vr_timeout_dialog),
        ], scroll=ft.ScrollMode.AUTO)

    def voice_pull_record(self, _):
        path = self.voice_path_input.value
        self.run_generic_cmd(f"pull {path} .")

    def voice_remove_record(self, _):
        path = self.voice_path_input.value
        self.run_generic_cmd(f"shell rm -rf {path}/*")

    def get_running_app_info(self, _):
        if not self.check_device(): return
        self.log("正在获取应用信息...")
        def task():
            # 1. 获取前台应用信息 (匹配 adb_device_utils.py 逻辑)
            res = adb_utils.run_adb_command("shell dumpsys activity top", self.selected_device)
            focus_info = None
            if res.returncode == 0 and res.stdout:
                for line in res.stdout.split('\n'):
                    if 'ACTIVITY' in line:
                        focus_info = line.strip()
                        break
            
            if not focus_info:
                res = adb_utils.run_adb_command("shell dumpsys window windows", self.selected_device)
                if res.returncode == 0 and res.stdout:
                    for line in res.stdout.split('\n'):
                        if 'mCurrentFocus' in line or 'mFocusedApp' in line:
                            focus_info = line.strip()
                            break
            
            if not focus_info:
                return self.log("无法获取前台应用信息")
            
            # 解析包名
            package_name = None
            activity_match = re.search(r'ACTIVITY\s+(\S+)', focus_info)
            if activity_match:
                full_name = activity_match.group(1)
                package_name = full_name.split('/')[0] if '/' in full_name else full_name
            
            if not package_name:
                focus_match = re.search(r'\{[^}]*\s+(\S+)/(\S*)\}', focus_info)
                if focus_match:
                    package_name = focus_match.group(1)
            
            if not package_name:
                return self.log("无法解析前台应用信息")
            
            self.log(f"包名: {package_name}")
            
            # 2. 获取版本信息
            res = adb_utils.run_adb_command(f"shell dumpsys package {package_name}", self.selected_device)
            if res.returncode == 0:
                version_match = re.search(r'versionName=(\S+)', res.stdout)
                if version_match:
                    self.log(f"应用 {package_name} 版本号: {version_match.group(1)}")
                else:
                    self.log(f"应用 {package_name} 无法获取版本信息")
            else:
                self.log(f"应用 {package_name} 不存在或无法访问")
                
        threading.Thread(target=task).start()

    def pick_install_dir(self, _):
        path = self._tk_pick_dir("选择 APK 目录")
        if not path:
            return
        self.log(f"准备从目录批量安装: {path}")
        self.batch_install_from_dir(path)

    def pick_verify_dir(self, _):
        path = self._tk_pick_dir("选择 APK 目录")
        if not path:
            return
        self.log(f"准备从目录验证版本: {path}")
        self.batch_verify_version_from_dir(path)

    def reinit_u2(self, _):
        if not self.check_device(): return
        self.log("正在重新初始化 uiautomator2...")
        def task():
            res = subprocess.run(["python", "-m", "uiautomator2", "init", "--serial", self.selected_device], capture_output=True, text=True)
            self.log(res.stdout or res.stderr)
            self.log("U2 初始化完成")
        threading.Thread(target=task).start()

    def show_version_check(self):
        self.version_check_result_view = ft.ListView(expand=True, height=300)
        self.releasenote_file_name_view = ft.Text("请选择文件...", color="grey500")
        self.main_container.content = ft.Column([
            ft.Row([
                self.create_action_button("选择集成清单", ft.Icons.FILE_OPEN, self.pick_releasenote_excel),
                self.releasenote_file_name_view,
            ], spacing=10),
            self.create_action_button("开始检查", ft.Icons.PLAY_ARROW, self.start_version_check),
            ft.Divider(),
            ft.Text("检查结果:", size=16, weight="bold"),
            self.version_check_result_view,
        ], scroll=ft.ScrollMode.AUTO)

    def pick_releasenote_excel(self, _):
        path = self._tk_pick_file("选择集成清单文件", [("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")])
        if not path:
            return
        self.excel_path = path
        self.log(f"已选择集成清单: {path}")
        if hasattr(self, 'releasenote_file_name_view'):
            self.releasenote_file_name_view.value = os.path.basename(path)
            self.releasenote_file_name_view.color = "blue"
            self.page.update()

    def show_settings(self):
        self.main_container.content = ft.Column([
            ft.Text("偏好设置", size=20, weight="bold"),
            ft.Divider(),
            ft.Switch(label="使用深色模式", value=self.page.theme_mode == ft.ThemeMode.DARK, on_change=self.toggle_theme),
            ft.Dropdown(label="界面语言", value="zh_CN", options=[
                ft.dropdown.Option("zh_CN", "简体中文"),
                ft.dropdown.Option("en_US", "English"),
            ]),
            ft.Slider(label="全局字体大小", min=10, max=20, value=12, divisions=10, on_change=self.change_font_size),
            ft.Divider(),
            ft.Text("系统工具", size=18, weight="bold"),
            ft.Row([
                self.create_action_button("打开日志目录", "folder_open", self.open_local_log_dir),
                self.create_action_button("检查软件更新", "system_update", lambda _: self.check_software_update(None, False)),
            ], spacing=10),
            ft.Row([
                self.create_action_button("查看操作历史", "history", self.show_history, color="orange400"),
                self.create_action_button("快速入门指南", "help_outline", self.show_guide, color="green400"),
            ], spacing=10),
            ft.Divider(),
            ft.Row([
                ft.Icon(ft.Icons.INFO_OUTLINE, color="grey500"),
                ft.Text(f"ADBTools Flet Edition v{config_manager.get_version()}", size=14, weight="bold"),
            ]),
            ft.Text("作者: 王克", size=12, color="grey500"),
            ft.TextButton("关于 ADBTools", on_click=self.show_about),
        ], scroll=ft.ScrollMode.AUTO)

    def show_about(self, _):
        dlg = ft.AlertDialog(
            title=ft.Text("关于 ADBTools"),
            content=ft.Column([
                ft.Text(f"版本: v{config_manager.get_version()}"),
                ft.Text("一个功能强大的 ADB 调试工具，支持多种设备管理功能。"),
                ft.Text("作者: 王克"),
                ft.Text("GitHub: https://github.com/wangke956/ADBTools"),
            ], tight=True),
            actions=[
                ft.TextButton("确定", on_click=lambda _: self.page.close(dlg)),
            ],
        )
        self.page.open(dlg)

    def show_guide(self, _):
        dlg = ft.AlertDialog(
            title=ft.Text("快速入门指南"),
            content=ft.Column([
                ft.Text("1. 连接 Android 设备并开启 USB 调试。"),
                ft.Text("2. 在上方下拉菜单中选择目标设备。"),
                ft.Text("3. 使用左侧导航栏切换不同功能模块。"),
                ft.Text("4. 应用管理支持拖拽安装和包名搜索。"),
                ft.Text("5. 专项功能包含项目特定的调试工具。"),
            ], tight=True, scroll=ft.ScrollMode.AUTO),
            actions=[
                ft.TextButton("我知道了", on_click=lambda _: self.page.close(dlg)),
            ],
        )
        self.page.open(dlg)

    def check_software_update(self, _, silent=False):
        if not silent:
            self.log("正在联网检查 ADBTools 更新...")
        current_version = config_manager.get_version()
        
        def task():
            try:
                import requests
                url = "https://api.github.com/repos/wangke956/ADBTools/releases/latest"
                headers = {'User-Agent': 'ADBTools-Flet-Update-Checker'}
                res = requests.get(url, headers=headers, timeout=10)
                
                if res.status_code == 200:
                    data = res.json()
                    latest_version = data.get('tag_name', '').replace('v', '').replace('V', '')
                    
                    if latest_version > current_version:
                        if not silent:
                            self.log(f"发现新版本: v{latest_version}")
                        
                        def open_url(_):
                            import webbrowser
                            webbrowser.open(data.get('html_url', ''))
                            self.page.close(dlg)

                        dlg = ft.AlertDialog(
                            title=ft.Text("发现新版本！"),
                            content=ft.Text(f"当前版本: v{current_version}\n最新版本: v{latest_version}\n\n是否前往 GitHub 下载更新？"),
                            actions=[
                                ft.TextButton("稍后再说", on_click=lambda _: self.page.close(dlg)),
                                ft.TextButton("前往下载", on_click=open_url),
                            ],
                        )
                        self.page.open(dlg)
                    else:
                        if not silent:
                            self.log(f"当前已是最新版本 (v{current_version})")
                else:
                    if not silent:
                        self.log(f"检查更新失败: HTTP {res.status_code}")
            except Exception as e:
                if not silent:
                    self.log(f"检查更新出错: {e}")
        
        threading.Thread(target=task).start()

    # --- 功能逻辑实现 ---

    def check_device(self):
        if not self.selected_device:
            self.page.snack_bar = ft.SnackBar(ft.Text("请先在上方选择一个设备！"))
            self.page.snack_bar.open = True
            self.page.update()
            return False
        return True

    def show_history(self, _):
        content = ft.ListView(expand=True, spacing=5, height=300)
        for h in self.operation_history:
            content.controls.append(ft.Text(h, size=12))
            
        dlg = ft.AlertDialog(
            title=ft.Text("最近操作历史"),
            content=ft.Container(content, width=500),
            actions=[ft.TextButton("关闭", on_click=lambda _: self.page.close(dlg))],
        )
        self.page.open(dlg)

    def change_font_size(self, e):
        size = int(e.control.value)
        self.page.theme = ft.Theme(text_theme=ft.TextTheme(body_medium=ft.TextStyle(size=size)))
        self.page.update()
        self.log(f"字体大小已调整为: {size}")

    def toggle_theme(self, e):
        self.page.theme_mode = ft.ThemeMode.DARK if e.control.value else ft.ThemeMode.LIGHT
        self.page.update()

    def run_generic_cmd(self, cmd):
        if not self.check_device(): return
        self.log(f"执行: adb -s {self.selected_device} {cmd}")
        def task():
            res = adb_utils.run_adb_command(cmd, self.selected_device)
            if res.stdout: self.log(res.stdout)
            if res.stderr: self.log(f"错误: {res.stderr}")
        threading.Thread(target=task).start()

    def take_screenshot(self, _):
        if not self.check_device(): return
        self.log("正在截取屏幕...")
        filename = f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        def task():
            res = adb_utils.get_screenshot(filename, self.selected_device)
            self.log(res)
        threading.Thread(target=task).start()

    def reboot_device(self, _):
        self.run_generic_cmd("reboot")

    def adb_root(self, _):
        self.run_generic_cmd("root")

    def get_device_info(self, _):
        if not self.check_device(): return
        self.log("正在获取设备详细信息...")
        def task():
            res1 = adb_utils.run_adb_command("shell getprop ro.product.model", self.selected_device)
            res2 = adb_utils.run_adb_command("shell getprop ro.build.version.release", self.selected_device)
            res3 = adb_utils.run_adb_command("shell getprop ro.build.version.sdk", self.selected_device)
            self.log(f"设备型号: {res1.stdout.strip()}")
            self.log(f"安卓版本: {res2.stdout.strip()}")
            self.log(f"SDK 版本: {res3.stdout.strip()}")
        threading.Thread(target=task).start()

    def on_file_result(self, e):
        if e.files:
            if hasattr(self, 'file_picker_mode') and self.file_picker_mode == "upload_to_device":
                local_paths = [f.path for f in e.files]
                # 获取当前的 FileBrowser 实例
                if isinstance(self.main_container.content, ft.Column) and len(self.main_container.content.controls) > 1:
                    fb = self.main_container.content.controls[1]
                    if isinstance(fb, FileBrowser):
                        fb.upload_files(local_paths)
                return

            file_path = e.files[0].path
            if file_path.endswith(".apk"):
                if hasattr(self, 'file_picker_mode') and self.file_picker_mode == "get_pkg_name":
                    self.log(f"正在获取 APK 包名: {file_path}")
                    def task():
                        pkg = adb_utils.aapt_get_package_name(file_path)
                        self.log(f"识别到包名: {pkg}")
                        if "失败" not in pkg:
                            self.pkg_input.current.value = pkg
                            self.page.update()
                    threading.Thread(target=task).start()
                else:
                    self.log(f"准备安装 APK: {file_path}")
                    self.run_generic_cmd(f"install -r \"{file_path}\"")
            elif file_path.lower().endswith((".xlsx", ".xls")):
                self.excel_path = file_path
                self.log(f"已选择集成清单: {file_path}")
                if hasattr(self, 'releasenote_file_name_view'):
                    self.releasenote_file_name_view.value = os.path.basename(file_path)
                    self.releasenote_file_name_view.color = "blue"
                    self.page.update()
        elif e.path: # 文件夹选择结果
            if self.file_picker_mode == "install":
                self.log(f"准备从目录批量安装: {e.path}")
                self.batch_install_from_dir(e.path)
            elif self.file_picker_mode == "verify":
                self.log(f"准备从目录验证版本: {e.path}")
                self.batch_verify_version_from_dir(e.path)
            elif self.file_picker_mode == "log_path":
                self.log(f"已选择日志保存路径: {e.path}")
                if hasattr(self, 'log_path_input'):
                    self.log_path_input.value = e.path
                    self.page.update()
            # 重置模式
            self.file_picker_mode = "install"

    def batch_verify_version_from_dir(self, folder_path):
        if not self.check_device(): return
        self.log(f"正在扫描目录中的 APK 并验证版本: {folder_path}")
        
        def task():
            try:
                apk_files = [f for f in os.listdir(folder_path) if f.endswith(".apk")]
                total = len(apk_files)
                if total == 0:
                    return self.log("未找到 APK 文件")
                
                self.log(f"找到 {total} 个 APK")
                match_count = 0
                mismatch_count = 0
                
                for index, apk in enumerate(apk_files):
                    full_path = os.path.join(folder_path, apk)
                    # 1. 获取本地 APK 版本和包名 (使用 aapt)
                    import subprocess
                    aapt_path = adb_utils.get_adb_path().replace("adb.exe", "aapt.exe").replace("adb", "aapt")
                    res = subprocess.run([aapt_path, "dump", "badging", full_path], capture_output=True, text=True, encoding='utf-8', errors='ignore')
                    
                    if res.returncode != 0:
                        self.log(f"[{index+1}/{total}] {apk}: aapt 解析失败")
                        continue
                    
                    pkg_match = re.search(r"package: name='([^']+)'", res.stdout)
                    ver_match = re.search(r"versionName='([^']+)'", res.stdout)
                    
                    if not pkg_match or not ver_match:
                        self.log(f"[{index+1}/{total}] {apk}: 无法解析包名或版本")
                        continue
                        
                    pkg = pkg_match.group(1)
                    local_ver = ver_match.group(1)
                    
                    # 2. 获取设备内版本
                    dev_res = adb_utils.run_adb_command(f"shell dumpsys package {pkg}", self.selected_device)
                    device_ver = "未安装"
                    if dev_res.returncode == 0 and dev_res.stdout:
                        match = re.search(r"versionName=([^\s]+)", dev_res.stdout)
                        if match: device_ver = match.group(1)
                    
                    if local_ver == device_ver:
                        self.log(f"[{index+1}/{total}] {pkg}: 版本一致 ({local_ver})")
                        match_count += 1
                    else:
                        self.log(f"[{index+1}/{total}] {pkg}: 版本不一致! 本地: {local_ver} | 设备: {device_ver}")
                        mismatch_count += 1
                
                self.log(f"版本验证完成! 一致: {match_count}, 不一致: {mismatch_count}")
            except Exception as e:
                self.log(f"版本验证发生异常: {e}")
        
        threading.Thread(target=task).start()

    def batch_install_from_dir(self, folder_path):
        if not self.check_device(): return
        self.log(f"正在扫描目录中的 APK: {folder_path}")
        
        def task():
            try:
                apk_files = [f for f in os.listdir(folder_path) if f.endswith(".apk")]
                total = len(apk_files)
                if total == 0:
                    return self.log("未找到 APK 文件")
                
                self.log(f"找到 {total} 个 APK")
                success_count = 0
                fail_count = 0
                
                for index, apk in enumerate(apk_files):
                    full_path = os.path.join(folder_path, apk)
                    self.log(f"[{index+1}/{total}] 正在处理: {apk}")
                    
                    # 1. 获取包名 (使用 aapt)
                    package_name = adb_utils.aapt_get_package_name(full_path)
                    if "失败" in package_name:
                        self.log(f"获取包名失败: {package_name}")
                        fail_count += 1
                        continue
                    
                    self.log(f"  识别到包名: {package_name}")
                    
                    # 2. 检查是否为特殊包
                    special_config = self.special_packages_config.get(package_name)
                    
                    if special_config:
                        self.log(f"  检测到特殊包配置: {special_config.get('description', '无描述')}")
                        # 获取安装路径
                        res = adb_utils.run_adb_command(f"shell pm path {package_name}", self.selected_device)
                        if res.returncode == 0 and "package:" in res.stdout:
                            device_apk_path = res.stdout.replace("package:", "").strip()
                            target_filename = os.path.basename(device_apk_path)
                            target_dir = os.path.dirname(device_apk_path)
                            target_full_path = f"{target_dir}/{target_filename}"
                            
                            # 是否删除
                            if special_config.get("delete_before_push", False):
                                self.log(f"  执行删除操作: {target_full_path}")
                                adb_utils.run_adb_command(f"shell rm -f {target_full_path}", self.selected_device)
                            
                            # 推送文件
                            self.log(f"  执行 Push 操作: {target_full_path}")
                            push_res = adb_utils.run_adb_command(f"push \"{full_path}\" {target_full_path}", self.selected_device)
                            if push_res.returncode == 0:
                                self.log(f"  Push 成功!")
                                success_count += 1
                            else:
                                self.log(f"  Push 失败: {push_res.stderr}")
                                fail_count += 1
                        else:
                            self.log(f"  未在设备上找到已安装路径，尝试普通安装...")
                            install_res = adb_utils.run_adb_command(f"install -r \"{full_path}\"", self.selected_device)
                            if install_res.returncode == 0:
                                success_count += 1
                            else:
                                fail_count += 1
                    else:
                        # 普通安装
                        self.log(f"  执行普通安装 (-r)...")
                        install_res = adb_utils.run_adb_command(f"install -r \"{full_path}\"", self.selected_device)
                        if install_res.returncode == 0:
                            self.log(f"  安装成功!")
                            success_count += 1
                        else:
                            self.log(f"  安装失败: {install_res.stderr}")
                            fail_count += 1
                
                self.log(f"批量安装完成! 成功: {success_count}, 失败: {fail_count}")
            except Exception as e:
                self.log(f"批量安装发生异常: {e}")
        
        threading.Thread(target=task).start()

    def get_current_app(self, _):
        if not self.check_device(): return
        def task():
            if self.connection_mode == "u2" and self.d:
                try:
                    curr = self.d.app_current()
                    pkg = curr.get('package')
                    act = curr.get('activity')
                    self.log(f"当前焦点 (U2): {pkg}/{act}")
                    self.pkg_input.current.value = pkg
                    self.page.update()
                    return
                except Exception as e:
                    self.log(f"U2 获取焦点失败: {e}, 切换 ADB 模式获取...")

            res = adb_utils.run_adb_command("shell dumpsys window | grep mCurrentFocus", self.selected_device)
            output = res.stdout.strip()
            self.log(f"当前焦点 (ADB): {output}")
            match = re.search(r'([a-zA-Z0-9._]+)/', output)
            if match:
                pkg = match.group(1)
                self.pkg_input.current.value = pkg
                self.page.update()
        threading.Thread(target=task).start()

    def show_package_input_dialog(self, title, callback):
        pkg_field = ft.TextField(
            label="应用包名",
            value=self.pkg_input.current.value if self.pkg_input.current else "",
            autofocus=True,
            on_submit=lambda _: close_dlg(True)
        )
        
        def close_dlg(confirmed):
            if confirmed and pkg_field.value:
                if self.pkg_input.current:
                    self.pkg_input.current.value = pkg_field.value
                callback(pkg_field.value)
            self.page.close(dlg)
            self.page.update()

        dlg = ft.AlertDialog(
            title=ft.Text(title),
            content=pkg_field,
            actions=[
                ft.TextButton("取消", on_click=lambda _: close_dlg(False)),
                ft.TextButton("确定", on_click=lambda _: close_dlg(True)),
            ],
        )
        self.page.open(dlg)
        self.page.update()

    def view_app_path(self, _):
        def action(pkg):
            self.run_generic_cmd(f"shell pm path {pkg}")
        
        pkg = self.pkg_input.current.value if self.pkg_input.current else ""
        if pkg: action(pkg)
        else: self.show_package_input_dialog("查看应用路径", action)

    def app_force_stop(self, _):
        def action(pkg):
            if self.connection_mode == "u2" and self.d:
                self.log(f"U2 执行: 停止应用 {pkg}")
                self.d.app_stop(pkg)
            else:
                self.run_generic_cmd(f"shell am force-stop {pkg}")
        
        self.show_package_input_dialog("强制停止应用", action)

    def app_clear_data(self, _):
        def action(pkg):
            if self.connection_mode == "u2" and self.d:
                self.log(f"U2 执行: 清理应用缓存 {pkg}")
                self.d.app_clear(pkg)
            else:
                self.run_generic_cmd(f"shell pm clear {pkg}")
        
        self.show_package_input_dialog("清理应用数据", action)

    def app_start(self, _):
        def action(pkg):
            if self.connection_mode == "u2" and self.d:
                self.log(f"U2 执行: 启动应用 {pkg}")
                self.d.app_start(pkg)
            else:
                self.run_generic_cmd(f"shell monkey -p {pkg} -c android.intent.category.LAUNCHER 1")
        
        self.show_package_input_dialog("启动应用", action)

    def app_uninstall(self, _):
        def action(pkg):
            if self.connection_mode == "u2" and self.d:
                self.log(f"U2 执行: 卸载应用 {pkg}")
                self.d.app_uninstall(pkg)
            else:
                self.run_generic_cmd(f"uninstall {pkg}")
        
        self.show_package_input_dialog("卸载应用", action)

    def open_local_log_dir(self, _):
        save_dir = self.log_path_input.value if hasattr(self, 'log_path_input') else os.getcwd()
        if sys.platform == 'win32':
            os.startfile(save_dir)
        else:
            self.log(f"当前目录: {save_dir}")

    def voice_start_record(self, _):
        self.run_generic_cmd("shell am broadcast -a com.saicmotor.voiceservice.START_RECORD")

    def voice_stop_record(self, _):
        self.run_generic_cmd("shell am broadcast -a com.saicmotor.voiceservice.STOP_RECORD")

    def sync_device_time(self, _):
        curr_time = datetime.now().strftime("%m%d%H%M%Y.%S")
        self.run_generic_cmd(f"shell date {curr_time}")

    def start_version_check(self, _):
        if not self.check_device(): return
        if not self.excel_path: return self.log("请先选择 Excel 清单文件")
        
        self.log(f"正在读取集成清单: {os.path.basename(self.excel_path)}")
        self.version_check_result_view.controls.clear()
        self.version_check_result_view.update()
        
        def task():
            try:
                import openpyxl
                # 使用 data_only=True 获取公式计算后的值
                wb = openpyxl.load_workbook(self.excel_path, data_only=True)
                if 'checkversion' not in wb.sheetnames:
                    return self.log("错误: Excel 文件中未找到 'checkversion' 工作表")
                
                ws = wb['checkversion']
                
                # 校验 B8 单元格 (匹配 AppVersionCheckThread 逻辑)
                b8_value = ws['B8'].value
                if b8_value != "packageName":
                    return self.log(f"错误: B8 单元格内容为 '{b8_value}'，预期为 'packageName'")
                
                self.log("校验通过，开始获取包名列表...")
                
                # 从 B9 开始循环 (匹配 AppVersionCheckThread 逻辑)
                check_items = []
                for i in range(9, 100):
                    pkg = ws.cell(row=i, column=2).value # B列
                    target_ver = ws.cell(row=i, column=4).value # D列
                    if not pkg: break
                    check_items.append((pkg, target_ver))
                
                total = len(check_items)
                self.log(f"共发现 {total} 个待检查项目")
                
                for index, (pkg, target_ver) in enumerate(check_items):
                    # 获取设备内版本
                    res = adb_utils.run_adb_command(f"shell dumpsys package {pkg}", self.selected_device)
                    current_ver = "未安装"
                    if res.returncode == 0 and res.stdout:
                        # 匹配 versionName=...
                        match = re.search(r"versionName=([^\s]+)", res.stdout)
                        if match: 
                            current_ver = match.group(1)
                    
                    status = "一致" if str(current_ver) == str(target_ver) else "不一致"
                    color = "green" if status == "一致" else "red"
                    
                    self.version_check_result_view.controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(f"[{index+1}/{total}] ", size=12, color="grey"),
                                    ft.Text(f"{pkg}", weight="bold", expand=True),
                                    ft.Text(status, color=color, weight="bold"),
                                ]),
                                ft.Row([
                                    ft.Text(f"预期: {target_ver}", size=12, color="blue"),
                                    ft.Text(f"实际: {current_ver}", size=12, color=color),
                                ], spacing=20),
                            ], spacing=2),
                            padding=5,
                            border=ft.border.all(1, "grey800"),
                            border_radius=5,
                            margin=ft.margin.only(bottom=5)
                        )
                    )
                    self.version_check_result_view.update()
                
                self.log("集成版本检查完成")
            except Exception as e:
                self.log(f"版本检查失败: {e}")
        
        threading.Thread(target=task).start()

    def show_input_dialog(self, _):
        def close_dlg(e):
            self.page.close(dlg)
            if e.control.text == "发送":
                txt = self.text_input.current.value
                self.run_generic_cmd(f"shell input text \"{txt}\"")

        dlg = ft.AlertDialog(
            title=ft.Text("输入文本到设备"),
            content=ft.TextField(ref=self.text_input, label="文本内容", autofocus=True),
            actions=[
                ft.TextButton("取消", on_click=close_dlg),
                ft.TextButton("发送", on_click=close_dlg),
            ],
        )
        self.page.open(dlg)

    def show_vr_timeout_dialog(self, _):
        def close_dlg(e):
            self.page.close(dlg)
            if e.control.text == "确认":
                val = self.text_input.current.value
                self.run_generic_cmd(f"shell setprop persist.sys.vr_timeout {val}")

        dlg = ft.AlertDialog(
            title=ft.Text("设置 VR 超时时间"),
            content=ft.TextField(ref=self.text_input, label="超时时间 (秒)", keyboard_type=ft.KeyboardType.NUMBER),
            actions=[
                ft.TextButton("取消", on_click=close_dlg),
                ft.TextButton("确认", on_click=close_dlg),
            ],
        )
        self.page.open(dlg)

def main(page: ft.Page):
    ADBToolsApp(page)

if __name__ == "__main__":
    ft.run(main)
