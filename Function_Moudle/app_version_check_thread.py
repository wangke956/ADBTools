import time
import os
from typing import Dict, Optional, Tuple, List
from collections import defaultdict

from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


class AppVersionCheckThread(QThread):
    """
    应用版本检查线程
    功能：读取Excel版本清单，对比设备端已安装应用的版本号
    """
    progress_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)
    release_note_signal = pyqtSignal(dict)
    version_compare_signal = pyqtSignal(dict)
    detailed_report_signal = pyqtSignal(dict)
    mismatch_list_signal = pyqtSignal(list)
    comparison_complete_signal = pyqtSignal(dict)

    def __init__(self, d, releasenote_file: str):
        super().__init__()
        self.d = d
        self.releasenote_file = releasenote_file
        self.release_note_dict: Dict[str, str] = {}
        self.compare_result_dict: Dict[str, Tuple[str, str, bool]] = {}
        self.mismatch_packages: List[Tuple[str, str, str]] = []
        self.summary_stats = {
            'total': 0,
            'matched': 0,
            'mismatched': 0,
            'not_found': 0
        }

    def _validate_file(self) -> bool:
        """校验版本清单文件是否有效"""
        self.progress_signal.emit("正在校验版本清单文件...")
        if not os.path.exists(self.releasenote_file):
            self.error_signal.emit(f"❌ 版本清单文件不存在：{self.releasenote_file}")
            return False
        if not self.releasenote_file.endswith(('.xlsx', '.xlsm')):
            self.error_signal.emit(f"❌ 文件格式错误，仅支持.xlsx/.xlsm：{self.releasenote_file}")
            return False
        return True

    def _read_excel_version(self) -> bool:
        """读取Excel中的版本清单"""
        self.progress_signal.emit("正在读取版本清单Excel文件...")
        try:
            wb = openpyxl.load_workbook(self.releasenote_file, read_only=True, data_only=True)

            if 'checkversion' not in wb.sheetnames:
                self.error_signal.emit("❌ Excel中不存在名为'checkversion'的工作表")
                wb.close()
                return False

            ws = wb['checkversion']
            cell_b8 = ws['B8'].value
            if cell_b8 != "packageName":
                self.error_signal.emit(f"❌ B8单元格值错误（预期：packageName，实际：{cell_b8}）")
                wb.close()
                return False

            self.progress_signal.emit("正在解析包名和版本号...")
            row_num = 9
            packages_found = 0

            while True:
                packagename_cell = ws.cell(row=row_num, column=2)
                version_cell = ws.cell(row=row_num, column=4)

                packagename = packagename_cell.value
                version = version_cell.value

                if packagename is None or str(packagename).strip() == "":
                    break

                clean_package = str(packagename).strip()
                clean_version = str(version).strip() if version and str(version).strip() != "" else "未知版本"

                self.release_note_dict[clean_package] = clean_version
                packages_found += 1

                if packages_found <= 5:
                    self.progress_signal.emit(f"  读取: {clean_package} -> {clean_version}")

                row_num += 1

                if row_num > 1000:
                    self.progress_signal.emit("⚠️ 警告：已读取1000行，停止读取")
                    break

            wb.close()

            if not self.release_note_dict:
                self.error_signal.emit("❌ Excel中未读取到任何有效的包名和版本号")
                return False

            self.progress_signal.emit(f"✅ 成功读取到 {len(self.release_note_dict)} 个应用的版本清单")
            if packages_found > 5:
                self.progress_signal.emit(f"  ... 还有 {packages_found - 5} 个应用")

            self.release_note_signal.emit(self.release_note_dict)
            return True

        except Exception as e:
            self.error_signal.emit(f"❌ 读取Excel失败：{str(e)}")
            return False

    def _get_device_app_version(self, packagename: str) -> Optional[str]:
        """获取设备端已安装应用的版本号 - 修复版本键名问题"""
        try:
            try:
                app_info = self.d.app_info(packagename)

                # 调试输出
                self.progress_signal.emit(f"  [DEBUG] 获取 {packagename} 的 app_info: {app_info}")

                if app_info:
                    # 尝试多种可能的键名
                    version_keys = ['versionName', 'version_name', 'version', 'VersionName']
                    for key in version_keys:
                        if key in app_info:
                            version = app_info.get(key)
                            if version:
                                return str(version)

                return None
            except ValueError as e:
                # 处理日期时间格式解析错误（如阿拉伯数字日期）
                if "does not match format" in str(e) or "time data" in str(e):
                    self.progress_signal.emit(f"  [WARNING] {packagename} 遇到日期格式问题，尝试备用方法")
                    # 尝试使用ADB命令获取版本信息
                    from Function_Moudle.adb_device_utils import get_app_version
                    device_id = getattr(self.d, 'serial', None)
                    if device_id:
                        success, version_info = get_app_version(device_id, packagename)
                        if success:
                            return version_info
                    return None
                else:
                    raise

        except Exception as e:
            self.progress_signal.emit(f"  [DEBUG] 获取 {packagename} 版本异常: {str(e)[:50]}")
            return None

    def _compare_version(self):
        """对比Excel版本和设备端版本 - 修复版"""
        self.progress_signal.emit("开始对比设备端应用版本...")

        # 重置统计
        self.summary_stats = {'total': 0, 'matched': 0, 'mismatched': 0, 'not_found': 0}
        self.mismatch_packages = []

        total_packages = len(self.release_note_dict)
        current_index = 0

        for packagename, excel_version in self.release_note_dict.items():
            current_index += 1
            time.sleep(0.1)

            progress_percent = int((current_index / total_packages) * 100)
            self.progress_signal.emit(
                f"[{progress_percent}%] 检查应用 ({current_index}/{total_packages}): {packagename}")

            # 获取设备端版本
            device_version = self._get_device_app_version(packagename)

            # 记录总数
            self.summary_stats['total'] += 1

            # 判定版本是否一致
            if device_version is None:
                is_match = False
                device_display = "❌ 未找到"
                self.summary_stats['not_found'] += 1
                self.progress_signal.emit(f"  ❌ 未找到版本: {packagename}")
            else:
                # 清理版本字符串
                excel_clean = str(excel_version).strip()
                device_clean = str(device_version).strip()

                # 调试输出
                self.progress_signal.emit(f"  [DEBUG] 表格版本: '{excel_clean}'，设备版本: '{device_clean}'")

                # 判断是否匹配
                if excel_clean == device_clean:
                    is_match = True
                    device_display = f"✅ {device_clean}"
                    self.summary_stats['matched'] += 1
                    self.progress_signal.emit(f"  ✅ 匹配: {packagename} = {device_clean}")
                else:
                    is_match = False
                    device_display = f"⚠️ {device_clean}"
                    self.summary_stats['mismatched'] += 1
                    # 记录版本不一致的包
                    self.mismatch_packages.append((packagename, excel_clean, device_clean))
                    self.progress_signal.emit(f"  ⚠️ 不一致: {packagename} 表格={excel_clean}, 设备={device_clean}")

            # 存储对比结果
            self.compare_result_dict[packagename] = (excel_version, device_display, is_match)

        # 发送信号
        self.version_compare_signal.emit(self.compare_result_dict)

        # 生成并发送报告
        self._generate_and_send_reports()

    def _generate_and_send_reports(self):
        """生成并发送各种报告"""
        # 验证统计
        total_calculated = (
                self.summary_stats.get('matched', 0) +
                self.summary_stats.get('mismatched', 0) +
                self.summary_stats.get('not_found', 0)
        )

        if total_calculated != self.summary_stats.get('total', 0):
            self.progress_signal.emit(f"⚠️ 统计不一致! 修正前: 总数={self.summary_stats.get('total', 0)}")
            self.summary_stats['total'] = total_calculated

        # 生成详细报告
        detailed_report = {
            "summary": dict(self.summary_stats),
            "all_apps": [],
            "mismatched_apps": [],
            "not_found_apps": []
        }

        for packagename, (excel_ver, device_ver_display, is_match) in self.compare_result_dict.items():
            device_ver_clean = device_ver_display.replace("✅ ", "").replace("⚠️ ", "").replace("❌ ", "")

            app_info = {
                "package_name": packagename,
                "excel_version": excel_ver,
                "device_version": device_ver_clean,
                "status": "matched" if is_match else "not_found" if "未找到" in device_ver_display else "mismatched"
            }

            detailed_report["all_apps"].append(app_info)

            if not is_match:
                if "未找到" in device_ver_display:
                    detailed_report["not_found_apps"].append(app_info)
                else:
                    detailed_report["mismatched_apps"].append(app_info)

        # 发送所有信号
        self.detailed_report_signal.emit(detailed_report)

        mismatch_package_names = [p[0] for p in self.mismatch_packages]
        self.mismatch_list_signal.emit(mismatch_package_names)
        self.comparison_complete_signal.emit(self.summary_stats)

        # 显示最终报告
        self._show_final_report()

    def _show_final_report(self):
        """显示最终报告"""
        if not self.mismatch_packages and self.summary_stats['not_found'] == 0:
            self.progress_signal.emit("\n" + "=" * 60)
            self.progress_signal.emit("✅ 恭喜！所有应用版本都与表格一致！")
            self.progress_signal.emit("=" * 60)
        elif self.mismatch_packages:
            self.progress_signal.emit("\n" + "=" * 60)
            self.progress_signal.emit(f"⚠️ 发现 {len(self.mismatch_packages)} 个版本不一致的应用")
            self.progress_signal.emit("=" * 60)

            for idx, (package, excel_ver, device_ver) in enumerate(self.mismatch_packages, 1):
                self.progress_signal.emit(f"{idx}. {package}")
                self.progress_signal.emit(f"   表格版本: {excel_ver}")
                self.progress_signal.emit(f"   设备版本: {device_ver}")
                self.progress_signal.emit("")
        else:
            self.progress_signal.emit("\n" + "=" * 60)
            self.progress_signal.emit(f"❌ 有 {self.summary_stats['not_found']} 个应用未找到")
            self.progress_signal.emit("=" * 60)

        # 最终统计
        self.progress_signal.emit("\n" + "=" * 60)
        self.progress_signal.emit("🎉 版本检查完成！最终统计：")
        self.progress_signal.emit("=" * 60)
        self.progress_signal.emit(f"📊 应用总数: {self.summary_stats['total']}")
        self.progress_signal.emit(f"✅ 版本一致: {self.summary_stats['matched']}")
        self.progress_signal.emit(f"⚠️ 版本不一致: {self.summary_stats['mismatched']}")
        self.progress_signal.emit(f"❌ 未找到: {self.summary_stats['not_found']}")
        self.progress_signal.emit("=" * 60)

    def run(self):
        """线程核心执行逻辑"""
        try:
            # 步骤1：校验文件
            if not self._validate_file():
                return

            # 步骤2：读取Excel版本清单
            if not self._read_excel_version():
                return

            # 步骤3：对比设备端版本
            self._compare_version()

        except Exception as e:
            import traceback
            self.error_signal.emit(f"❌ 线程执行异常：{str(e)}")
            self.error_signal.emit(f"详细错误信息：\n{traceback.format_exc()}")


# -------------------------- 快速测试版本 --------------------------
def quick_test():
    """快速测试版本获取"""
    import uiautomator2 as u2

    try:
        d = u2.connect("192.168.43.47:5555")
        print(f"✅ 已连接设备: {d.info}")
    except Exception as e:
        print(f"❌ 连接失败: {e}")
        return

    # 测试几个包
    test_packages = [
        "com.saicmotor.voiceservice",
        "com.xtad.lbsadapter",
        "com.saicmotor.voicetts"
    ]

    for package in test_packages:
        try:
            app_info = d.app_info(package)
            print(f"\n{package}:")
            print(f"  app_info: {app_info}")
            if app_info:
                print(f"  所有键: {list(app_info.keys())}")
                for key in app_info:
                    print(f"    {key}: {app_info[key]}")
        except Exception as e:
            print(f"\n{package}: ❌ 异常: {e}")


if __name__ == "__main__":
    # 先运行快速测试
    print("=== 快速测试 ===")
    quick_test()
    print("\n=== 快速测试结束 ===\n")

    # 如果需要完整测试，取消下面的注释
    # import sys
    # from PyQt5.QtWidgets import QApplication
    #
    # app = QApplication(sys.argv)
    # d = u2.connect("192.168.43.47:5555")
    # excel_file = "版本清单.xlsx"
    #
    # check_thread = AppVersionCheckThread(d, excel_file)
    # check_thread.start()
    # check_thread.wait()